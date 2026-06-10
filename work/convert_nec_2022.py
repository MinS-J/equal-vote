from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pandas as pd

from paths import DATA_DIR, INPUTS_DIR

SOURCE = (
    INPUTS_DIR
    / "지방선거"
    / "중앙선거관리위원회_제8회 전국동시지방선거 개표결과_20220601.xlsx"
)
OUT_DIR = DATA_DIR

REGULAR_GUBUN = {"소계", "관내사전투표", "선거일투표"}
SPECIAL_EUPMYEONDONG = {
    "",
    "합계",
    "거소투표",
    "관외사전투표",
    "잘못 투입·구분된 투표지",
}


def clean_int(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).replace(",", "").strip()
    if text == "":
        return None
    try:
        return int(text)
    except ValueError:
        return None


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_sheet(wb, sheet_name: str):
    ws = wb[sheet_name]
    row1 = [cell.value for cell in ws[1]]
    row2 = [cell.value for cell in ws[2]]
    row3 = [cell.value for cell in ws[3]]

    candidate_cols = []
    for idx, value in enumerate(row2):
        label = clean_text(value)
        if label.startswith("후보") or label.startswith("정당"):
            candidate_cols.append(idx)

    candidate_names = {
        f"candidate_{i + 1}": clean_text(row3[col])
        for i, col in enumerate(candidate_cols)
    }

    base_cols = {}
    for idx, value in enumerate(row1):
        text = clean_text(value)
        if text in {"선거구명", "시도명"}:
            base_cols["sido_or_district"] = idx
        elif text in {"구시군명", "구시군"}:
            base_cols["sigungu"] = idx
        elif text in {"선거구(구시군)", "선거구명"} and "district" not in base_cols:
            base_cols["district"] = idx
        elif text == "읍면동명":
            base_cols["eupmyeondong"] = idx
        elif text == "구분":
            base_cols["gubun"] = idx
        elif text == "선거인수":
            base_cols.setdefault("electors", idx)
        elif text == "투표수":
            base_cols["votes"] = idx
        elif text == "계":
            base_cols["valid_total"] = idx
        elif text == "무효투표수":
            base_cols["invalid"] = idx
        elif text == "기권수":
            base_cols["abstain"] = idx

    rows = []
    for excel_row in ws.iter_rows(min_row=4, values_only=True):
        electors = clean_int(excel_row[base_cols["electors"]]) if "electors" in base_cols else None
        votes = clean_int(excel_row[base_cols["votes"]]) if "votes" in base_cols else None
        if electors is None and votes is None:
            continue

        row = {
            "sheet": sheet_name,
            "sido_or_district": clean_text(excel_row[base_cols["sido_or_district"]])
            if "sido_or_district" in base_cols
            else "",
            "sigungu": clean_text(excel_row[base_cols["sigungu"]]) if "sigungu" in base_cols else "",
            "district": clean_text(excel_row[base_cols["district"]]) if "district" in base_cols else "",
            "eupmyeondong": clean_text(excel_row[base_cols["eupmyeondong"]])
            if "eupmyeondong" in base_cols
            else "",
            "gubun": clean_text(excel_row[base_cols["gubun"]]) if "gubun" in base_cols else "",
            "electors": electors,
            "votes": votes,
            "valid_total": clean_int(excel_row[base_cols["valid_total"]])
            if "valid_total" in base_cols
            else None,
            "invalid": clean_int(excel_row[base_cols["invalid"]]) if "invalid" in base_cols else None,
            "abstain": clean_int(excel_row[base_cols["abstain"]]) if "abstain" in base_cols else None,
        }

        candidate_values = []
        for i, col in enumerate(candidate_cols):
            value = clean_int(excel_row[col]) or 0
            row[f"candidate_{i + 1}"] = value
            candidate_values.append(value)

        row["candidate_1"] = row.get("candidate_1", 0)
        row["candidate_2"] = row.get("candidate_2", 0)
        row["other_candidates"] = sum(candidate_values[2:])
        row["candidate_count"] = len(candidate_cols)
        rows.append(row)

    return pd.DataFrame(rows), candidate_names


def write_csv_if_possible(frame: pd.DataFrame, path: Path):
    try:
        frame.to_csv(path, index=False, encoding="utf-8-sig")
        return str(path)
    except PermissionError:
        alt = path.with_name(f"{path.stem}_new{path.suffix}")
        frame.to_csv(alt, index=False, encoding="utf-8-sig")
        return str(alt)


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)

    frames = []
    metadata = {}
    for sheet_name in wb.sheetnames:
        frame, candidate_names = normalize_sheet(wb, sheet_name)
        frames.append(frame)
        metadata[sheet_name] = candidate_names
        write_csv_if_possible(frame, OUT_DIR / f"{sheet_name}.csv")

    combined = pd.concat(frames, ignore_index=True, sort=False)
    regular = combined[
        combined["gubun"].isin(REGULAR_GUBUN)
        & ~combined["eupmyeondong"].isin(SPECIAL_EUPMYEONDONG)
    ].copy()
    advance = regular[regular["gubun"] == "관내사전투표"].copy()

    combined.to_pickle(OUT_DIR / "nec_2022_normalized.pkl")
    normalized_csv = write_csv_if_possible(combined, OUT_DIR / "nec_2022_normalized.csv")
    regular.to_pickle(OUT_DIR / "nec_2022_regular_rows.pkl")
    regular_csv = write_csv_if_possible(regular, OUT_DIR / "nec_2022_regular_rows.csv")
    advance.to_pickle(OUT_DIR / "nec_2022_advance_rows.pkl")
    advance_csv = write_csv_if_possible(advance, OUT_DIR / "nec_2022_advance_rows.csv")
    (OUT_DIR / "candidate_metadata.json").write_text(
        json.dumps(metadata, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    summary = {
        "source": str(SOURCE),
        "rows": int(len(combined)),
        "sheets": {name: int(len(frame)) for name, frame in zip(wb.sheetnames, frames)},
        "outputs": [
            str(OUT_DIR / "nec_2022_normalized.pkl"),
            normalized_csv,
            str(OUT_DIR / "nec_2022_regular_rows.pkl"),
            regular_csv,
            str(OUT_DIR / "nec_2022_advance_rows.pkl"),
            advance_csv,
            str(OUT_DIR / "candidate_metadata.json"),
        ],
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
